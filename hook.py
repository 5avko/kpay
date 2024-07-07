import os
import datetime
import logging
import aiocryptopay
import pydantic
import asyncio
import re
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Text
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils import executor
from aiocryptopay import AioCryptoPay, Networks
import openpyxl
from openpyxl.styles import PatternFill, Alignment

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Логирование версий библиотек
logger.info(f"aiocryptopay version: {aiocryptopay.__version__}")
logger.info(f"pydantic version: {pydantic.__version__}")

TOKEN = '7266212298:AAHFblaG0mIPgNGVQcLDW3WawS0Qo7N7D9w'
CRYPTOBOT_TOKEN = '224616:AAhUpDVfMjwoLKjNQoZHjpWgIDuk5BknMsu'
CRYPTOBOT_USER_ID = 1189549622  # ID бота @CryptoBot
USER_FILES_DIR = '/Users/workacc/Documents/botforCryptoTeam/resourse/anoun'
os.makedirs(USER_FILES_DIR, exist_ok=True)

class UserState(StatesGroup):
    REGISTERED = State()
    WORKING_DAY_STARTED = State()
    WORKING_DAY_NOT_STARTED = State()
    AWAITING_PAYMENT = State()

bot = Bot(token=TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)

crypto = AioCryptoPay(token=CRYPTOBOT_TOKEN, network=Networks.MAIN_NET)
logger.info(f"Crypto object initialized: {crypto}")
user_subscriptions = {}
user_payments = {}

def check_subscription(user_id):
    return user_id in user_subscriptions and user_subscriptions[user_id] > datetime.datetime.now()

async def create_invoice(amount, currency):
    try:
        logger.info(f"Attempting to create invoice: amount={amount}, currency={currency}")
        invoice = await crypto.create_invoice(
            asset=currency,
            amount=amount,
            payload=f"subscription:{amount}"
        )
        logger.info(f"Invoice created: {invoice}")
        result = invoice.model_dump() if hasattr(invoice, 'model_dump') else invoice.dict()
        logger.info(f"Invoice data: {result}")
        return result
    except Exception as e:
        logger.error(f"Error creating invoice: {str(e)}", exc_info=True)
        return None

async def check_invoice_status(invoice_id):
    try:
        invoices = await crypto.get_invoices(invoice_ids=[invoice_id])
        if invoices and len(invoices) > 0:
            return invoices[0].status
        else:
            logger.error(f"No invoice found with id: {invoice_id}")
            return None
    except Exception as e:
        logger.error(f"Error checking invoice status: {str(e)}", exc_info=True)
        return None

async def check_crypto_connection():
    try:
        me = await crypto.get_me()
        logger.info(f"Connected to Crypto Bot: {me}")
    except Exception as e:
        logger.error(f"Failed to connect to Crypto Bot: {str(e)}", exc_info=True)

@dp.message_handler(commands=['start'])
async def start(message: types.Message):
    user_id = message.from_user.id
    if check_subscription(user_id):
        await show_main_menu(message)
    else:
        await show_subscription_options(message)

async def show_subscription_options(message: types.Message):
    keyboard = InlineKeyboardMarkup(row_width=2)
    keyboard.add(InlineKeyboardButton("1 месяц - 10$", callback_data="sub:1:month:10"),
                 InlineKeyboardButton("3 месяца - 25$", callback_data="sub:3:months:25"),
                 InlineKeyboardButton("1 год - 90$", callback_data="sub:1:year:90"))
    await message.reply("Выберите план подписки:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith('pay:'))
async def process_payment(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    parts = callback_query.data.split(':')
    
    logger.info(f"Callback data: {callback_query.data}")
    logger.info(f"Parsed parts: {parts}")
    
    if len(parts) != 5:
        await bot.answer_callback_query(callback_query.id, "Неверные данные для оплаты")
        return
    
    _, subscription_duration, subscription_period, price, currency = parts

    logger.info(f"Creating invoice for user {user_id}, price: {price} {currency}")
    invoice = await create_invoice(float(price), currency)
    if invoice and 'bot_invoice_url' in invoice:
        pay_url = invoice['bot_invoice_url']
        invoice_id = invoice['invoice_id']
        keyboard = InlineKeyboardMarkup().add(InlineKeyboardButton("Оплатить", url=pay_url))
        await bot.send_message(user_id, f"Для активации подписки на {subscription_duration} {subscription_period}, оплатите {price} {currency}", reply_markup=keyboard)
        
        # Store payment info
        user_payments[user_id] = {
            'invoice_id': invoice_id,
            'amount': float(price),
            'subscription_option': f"{subscription_duration}:{subscription_period}"
        }
                # Start checking payment status
        asyncio.create_task(check_payment_status(user_id, invoice_id, float(price), f"{subscription_duration}:{subscription_period}"))
    else:
        logger.error(f"Failed to create invoice for user {user_id}. Invoice data: {invoice}")
        await bot.send_message(user_id, "Ошибка при создании счета. Пожалуйста, попробуйте позже или обратитесь в поддержку.")

    await bot.answer_callback_query(callback_query.id)


async def check_payment_status(user_id, invoice_id, amount, subscription_option):
    for _ in range(60):  # Check for 10 minutes (60 * 10 seconds)
        status = await check_invoice_status(invoice_id)
        logger.info(f"Checking payment status for invoice {invoice_id}: {status}")
        if status == 'paid':
            await process_successful_payment(user_id, amount, subscription_option)
            return
        elif status == 'expired':
            await bot.send_message(user_id, "Срок действия инвойса истек. Пожалуйста, попробуйте снова.")
            return
        await asyncio.sleep(10)  # Wait for 10 seconds before checking again
    
    await bot.send_message(user_id, "Время ожидания оплаты истекло. Если вы уже оплатили, пожалуйста, свяжитесь с поддержкой.")

async def process_successful_payment(user_id, amount, subscription_option):
    duration, period = subscription_option.split(':')
    if period == 'month':
        days = int(duration) * 30
    elif period == 'months':
        days = int(duration) * 30
    elif period == 'year':
        days = 365
    else:
        await bot.send_message(user_id, "Ошибка в определении длительности подписки.")
        return

    expiration_date = datetime.datetime.now() + datetime.timedelta(days=days)
    user_subscriptions[user_id] = expiration_date
    
    await bot.send_message(user_id, f"Подписка активирована до {expiration_date.strftime('%d.%m.%Y')}")
    await show_main_menu(await bot.send_message(user_id, "Выберите действие:"))

@dp.callback_query_handler(lambda c: c.data.startswith('sub:'))
async def process_subscription(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    option = callback_query.data
    parts = option.split(':')
    
    if len(parts) != 4:
        await bot.answer_callback_query(callback_query.id, "Неверный выбор подписки")
        return
    
    _, duration, period, price = parts
    await bot.answer_callback_query(callback_query.id)
    await show_currency_options(callback_query.message, float(price), f"{duration}:{period}")

async def show_currency_options(message: types.Message, price: float, subscription_option: str):
    currencies = await get_available_currencies()
    keyboard = InlineKeyboardMarkup(row_width=2)
    for currency in currencies:
        converted_price = await convert_usd_to_crypto(price, currency)
        formatted_price = f"{converted_price:.8f}".rstrip('0').rstrip('.')
        keyboard.insert(InlineKeyboardButton(
            f"{currency} ({formatted_price})",
            callback_data=f"pay:{subscription_option}:{converted_price:.8f}:{currency}"
        ))
    await message.reply(f"Выберите валюту для оплаты (эквивалент {price}$):", reply_markup=keyboard)

async def convert_usd_to_crypto(usd_amount, crypto_currency):
    # Примерные курсы обмена (в реальном приложении здесь должен быть запрос к API)
    rates = {
        'BTC': 0.000020,  # 1 USD = 0.000025 BTC
        'ETH': 0.00035,   # 1 USD = 0.00050 ETH
        'USDT': 1,        # 1 USD = 1 USDT
        'TON': 0.14,         # 1 USD = 2 TON
        'TRX': 8,        # 1 USD = 15 TRX
        'NOT': 85,        # 1 USD = 10 NOT (примерный курс)
    }
    if crypto_currency in rates:
        return usd_amount * rates[crypto_currency]
    else:
        return usd_amount  # Если валюта неизвестна, возвращаем исходную сумму в USD

async def show_main_menu(message: types.Message):
    reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
        KeyboardButton('Начать рабочий день'),
        KeyboardButton('Скачать таблицу')
    )
    await message.reply("Выберите опцию из меню.", reply_markup=reply_keyboard)
    await UserState.WORKING_DAY_NOT_STARTED.set()

@dp.message_handler(Text(equals="Начать рабочий день"), state=UserState.WORKING_DAY_NOT_STARTED)
async def start_day(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not check_subscription(user_id):
        await show_subscription_options(message)
        return

    file_path = os.path.join(USER_FILES_DIR, f'user_{user_id}.xlsx')
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ['Дата', 'Время', 'ID сделки', 'Тип сделки', 'Монета', 'Количество', 'Цена за единицу', 'Сумма в RUB', 'Способ оплаты', 'Реквизиты', 'Примечание']
        sheet.append(headers)
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

    now = datetime.datetime.now()
    date = now.strftime('%d.%m.%Y')
    time = now.strftime('%H:%M')
    row = [date, time, '', 'Начало рабочего дня', '', '', '', '', '', '', '']
    sheet.append(row)

    for col in range(1, 12):
        sheet.cell(row=sheet.max_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=11)
    sheet.cell(row=sheet.max_row, column=1).fill = PatternFill(start_color='FFFF00', fill_type='solid')

    workbook.save(file_path)

    reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
        KeyboardButton('Закончить рабочий день'),
        KeyboardButton('Скачать таблицу')
    )
    await message.reply('Рабочий день начат. Теперь вы можете добавлять сделки.', reply_markup=reply_keyboard)
    await UserState.WORKING_DAY_STARTED.set()

async def get_available_currencies():
    try:
        currencies = await crypto.get_currencies()
        supported_currencies = ['BTC', 'ETH', 'USDT', 'TON', 'TRX', 'NOT']
        return [currency.code for currency in currencies if currency.is_blockchain and currency.code in supported_currencies]
    except Exception as e:
        logger.error(f"Error getting available currencies: {str(e)}", exc_info=True)
        return ['USDT', 'TON', 'TRX', 'NOT']  
@dp.message_handler(Text(equals="Закончить рабочий день"), state=UserState.WORKING_DAY_STARTED)
async def end_day(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if not check_subscription(user_id):
        await show_subscription_options(message)
        return

    file_path = os.path.join(USER_FILES_DIR, f'user_{user_id}.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    now = datetime.datetime.now()
    date = now.strftime('%d.%m.%Y')
    time = now.strftime('%H:%M')
    row = [date, time, '', 'Конец рабочего дня', '', '', '', '', '', '', '']
    sheet.append(row)

    for col in range(1, 12):
        sheet.cell(row=sheet.max_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=11)
    sheet.cell(row=sheet.max_row, column=1).fill = PatternFill(start_color='FF0000', fill_type='solid')

    workbook.save(file_path)

    reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
        KeyboardButton('Начать рабочий день'),
        KeyboardButton('Скачать таблицу')
    )
    await message.reply('Рабочий день завершен.', reply_markup=reply_keyboard)
    await UserState.WORKING_DAY_NOT_STARTED.set()

@dp.message_handler(Text(equals="Скачать таблицу"), state='*')
async def download_table(message: types.Message):
    user_id = message.from_user.id
    file_path = os.path.join(USER_FILES_DIR, f'user_{user_id}.xlsx')
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            await bot.send_document(message.chat.id, types.InputFile(file, filename=f'user_{user_id}.xlsx'))
    else:
        await message.reply('У вас еще нет данных.')

@dp.message_handler(lambda message: message.from_user.id == CRYPTOBOT_USER_ID and "оплатил(а) ваш счёт" in message.text)
async def handle_cryptobot_payment(message: types.Message):
    logger.info(f"Received payment notification: {message.text}")
    # Extract invoice ID from the message
    invoice_id_match = re.search(r'#IV(\d+)', message.text)
    if invoice_id_match:
        invoice_id = invoice_id_match.group(1)
        # Find the corresponding user and process the payment
        for user_id, user_data in user_payments.items():
            if user_data['invoice_id'] == invoice_id:
                await process_successful_payment(user_id, user_data['amount'], user_data['subscription_option'])
                del user_payments[user_id]
                return
    logger.error(f"Could not process payment notification: {message.text}")

async def on_startup(_):
    await check_crypto_connection()

if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)