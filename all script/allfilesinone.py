import os
import datetime
import logging
import aiocryptopay
import pydantic
import asyncio
import re
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Text
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils import executor
from aiocryptopay import AioCryptoPay, Networks
import signal

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
ADMIN_IDS = [5150619792, 5150619792]  # Замените на реальные ID администраторов
# Логирование версий библиотек
logger.info(f"aiocryptopay version: {aiocryptopay.__version__}")
logger.info(f"pydantic version: {pydantic.__version__}")


TOKEN = '7266212298:AAHFblaG0mIPgNGVQcLDW3WawS0Qo7N7D9w'
CRYPTOBOT_TOKEN = '224616:AAhUpDVfMjwoLKjNQoZHjpWgIDuk5BknMsu'
CRYPTOBOT_USER_ID = 1189549622  # ID бота @CryptoBot
USER_FILES_DIR = './all_script/'


#TODO: to check whether we need it
#os.makedirs(USER_FILES_DIR, exist_ok=True)

class UserState(StatesGroup):
    REGISTERED = State()
    WORKING_DAY_STARTED = State()
    WORKING_DAY_NOT_STARTED = State()
    AWAITING_PAYMENT = State()
    ADMIN_GRANTING_SUBSCRIPTION = State()
    ADMIN_GRANTING_FREE_SUBSCRIPTION = State()  # Новое состояние # Новое состояние

bot = Bot(token=TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)

crypto = AioCryptoPay(token=CRYPTOBOT_TOKEN, network=Networks.MAIN_NET)
logger.info(f"Crypto object initialized: {crypto}")
user_subscriptions = {}
user_payments = {}

@dp.message_handler(commands=['admin'], state='*')
async def admin_panel(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.reply("У вас нет доступа к этой команде.")
        return
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("Просмотр подписок", callback_data="admin_view_subscriptions"))
    keyboard.add(InlineKeyboardButton("Выдать подписку", callback_data="admin_grant_subscription"))
    
    await message.reply("Панель администратора:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data == 'admin_view_subscriptions')
async def view_subscriptions(callback_query: types.CallbackQuery):
    if callback_query.from_user.id not in ADMIN_IDS:
        await bot.answer_callback_query(callback_query.id, "У вас нет доступа к этой функции.")
        return

    subscriptions_text = "Активные подписки:\n\n"
    for user_id, expiration_date in user_subscriptions.items():
        subscriptions_text += f"Пользователь ID: {user_id}, Подписка до: {expiration_date.strftime('%d.%m.%Y')}\n"
    
    await bot.send_message(callback_query.from_user.id, subscriptions_text)
    await bot.answer_callback_query(callback_query.id)

@dp.callback_query_handler(lambda c: c.data == 'admin_grant_subscription')
async def grant_subscription_start(callback_query: types.CallbackQuery):
    if callback_query.from_user.id not in ADMIN_IDS:
        await bot.answer_callback_query(callback_query.id, "У вас нет доступа к этой функции.")
        return

    await bot.send_message(callback_query.from_user.id, "Введите ID пользователя и количество дней подписки в формате: ID дни")
    await bot.answer_callback_query(callback_query.id)
    await UserState.ADMIN_GRANTING_SUBSCRIPTION.set()

@dp.message_handler(state=UserState.ADMIN_GRANTING_SUBSCRIPTION)
async def grant_subscription(message: types.Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        await message.reply("У вас нет доступа к этой функции.")
        return

    try:
        user_id, days = map(int, message.text.split())
        expiration_date = datetime.datetime.now() + datetime.timedelta(days=days)
        user_subscriptions[user_id] = expiration_date
        save_subscriptions()
        await message.reply(f"Подписка выдана пользователю {user_id} до {expiration_date.strftime('%d.%m.%Y')}")
    except ValueError:
        await message.reply("Неверный формат. Используйте: ID дни")
    
    await state.finish()

async def check_subscriptions():
    current_time = datetime.datetime.now()
    for user_id, expiration_date in list(user_subscriptions.items()):
        if expiration_date < current_time:
            await bot.send_message(user_id, "Ваша подписка истекла. Пожалуйста, обновите подписку для продолжения использования бота.")
            del user_subscriptions[user_id]
            save_subscriptions()
        elif (expiration_date - current_time).days <= 3:
            await bot.send_message(user_id, f"Ваша подписка истекает через {(expiration_date - current_time).days} дней. Не забудьте продлить ее.")
        elif (expiration_date - current_time).days == 0:
            await bot.send_message(user_id, "Ваша подписка истекает сегодня. Пожалуйста, продлите ее для дальнейшего использования бота.")
            
def check_active_subscription(func):
    async def wrapper(message: types.Message, *args, **kwargs):
        user_id = message.from_user.id
        if user_id in ADMIN_IDS or check_subscription(user_id):
            return await func(message, *args, **kwargs)
        else:
            await show_subscription_options(message)
    return wrapper

def check_subscription(user_id):
    return user_id in user_subscriptions and user_subscriptions[user_id] > datetime.datetime.now()

async def create_invoice(amount, currency):
    try:
        logger.info(f"Attempting to create invoice: amount={amount}, currency={currency}")
        invoice = await crypto.create_invoice(
            asset=currency,
            amount=amount,
            payload=f"subscription:{amount}"
        )
        logger.info(f"Invoice created: {invoice}")
        result = invoice.model_dump() if hasattr(invoice, 'model_dump') else invoice.dict()
        logger.info(f"Invoice data: {result}")
        return result
    except Exception as e:
        logger.error(f"Error creating invoice: {str(e)}", exc_info=True)
        return None

async def check_invoice_status(invoice_id):
    try:
        invoices = await crypto.get_invoices(invoice_ids=[invoice_id])
        if invoices and len(invoices) > 0:
            return invoices[0].status
        else:
            logger.error(f"No invoice found with id: {invoice_id}")
            return None
    except Exception as e:
        logger.error(f"Error checking invoice status: {str(e)}", exc_info=True)
        return None

async def check_crypto_connection():
    try:
        me = await crypto.get_me()
        logger.info(f"Connected to Crypto Bot: {me}")
    except Exception as e:
        logger.error(f"Failed to connect to Crypto Bot: {str(e)}", exc_info=True)


async def schedule_subscription_check(dp: Dispatcher):
    while True:
        await check_subscriptions()
        await asyncio.sleep(86400)  # 24 часа
        
@dp.message_handler(commands=['start'])
async def start(message: types.Message):
    user_id = message.from_user.id
    if user_id in ADMIN_IDS:
        await admin_panel(message)
    elif check_subscription(user_id):
        await show_main_menu(message)
    else:
        await show_subscription_options(message)

async def show_subscription_options(message: types.Message):
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(InlineKeyboardButton("3 дня - 2 USDT", callback_data="sub:3:days:2"),
                 InlineKeyboardButton("5 дней - 3 USDT", callback_data="sub:5:days:3"),
                 InlineKeyboardButton("7 дней - 4 USDT", callback_data="sub:7:days:4"))
    await message.reply("Выберите план подписки:", reply_markup=keyboard)

@dp.callback_query_handler(lambda c: c.data.startswith('pay:'))
async def process_payment(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    parts = callback_query.data.split(':')
    
    logger.info(f"Callback data: {callback_query.data}")
    logger.info(f"Parsed parts: {parts}")
    
    if len(parts) != 5:
        await bot.answer_callback_query(callback_query.id, "Неверные данные для оплаты")
        return
    
    _, subscription_duration, subscription_period, price, currency = parts

    logger.info(f"Creating invoice for user {user_id}, price: {price} {currency}")
    invoice = await create_invoice(float(price), currency)
    if invoice and 'bot_invoice_url' in invoice:
        pay_url = invoice['bot_invoice_url']
        invoice_id = invoice['invoice_id']
        keyboard = InlineKeyboardMarkup().add(InlineKeyboardButton("Оплатить", url=pay_url))
        await bot.send_message(user_id, f"Для активации подписки на {subscription_duration} {subscription_period}, оплатите {price} {currency}", reply_markup=keyboard)
        
        user_payments[user_id] = {
            'invoice_id': invoice_id,
            'amount': float(price),
            'subscription_option': f"{subscription_duration}:{subscription_period}"
        }
        asyncio.create_task(check_payment_status(user_id, invoice_id, float(price), f"{subscription_duration}:{subscription_period}"))
    else:
        logger.error(f"Failed to create invoice for user {user_id}. Invoice data: {invoice}")
        await bot.send_message(user_id, "Ошибка при создании счета. Пожалуйста, попробуйте позже или обратитесь в поддержку.")

    await bot.answer_callback_query(callback_query.id)

async def check_payment_status(user_id, invoice_id, amount, subscription_option):
    for _ in range(60):  # Check for 10 minutes (60 * 10 seconds)
        status = await check_invoice_status(invoice_id)
        logger.info(f"Checking payment status for invoice {invoice_id}: {status}")
        if status == 'paid':
            await process_successful_payment(user_id, amount, subscription_option)
            return
        elif status == 'expired':
            await bot.send_message(user_id, "Срок действия инвойса истек. Пожалуйста, попробуйте снова.")
            return
        await asyncio.sleep(10)  # Wait for 10 seconds before checking again
    
    await bot.send_message(user_id, "Время ожидания оплаты истекло. Если вы уже оплатили, пожалуйста, свяжитесь с поддержкой.")

async def process_successful_payment(user_id, amount, subscription_option):
    duration, period = subscription_option.split(':')
    if period == 'month':
        days = int(duration) * 30
    elif period == 'months':
        days = int(duration) * 30
    elif period == 'year':
        days = 365
    else:
        await bot.send_message(user_id, "Ошибка в определении длительности подписки.")
        return

    expiration_date = datetime.datetime.now() + datetime.timedelta(days=days)
    user_subscriptions[user_id] = expiration_date
    save_subscriptions() 
    
    # Сохраняем информацию о подписке в файл
    with open('subscriptions.txt', 'a') as f:
        f.write(f"{user_id},{expiration_date.isoformat()}\n")
    
    await bot.send_message(user_id, f"Подписка активирована до {expiration_date.strftime('%d.%m.%Y')}")
    await show_main_menu(await bot.send_message(user_id, "Выберите действие:"))
    
@dp.callback_query_handler(lambda c: c.data.startswith('sub:'))
async def process_subscription(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id
    option = callback_query.data
    parts = option.split(':')
    
    if len(parts) != 4:
        await bot.answer_callback_query(callback_query.id, "Неверный выбор подписки")
        return
    
    _, duration, period, price = parts
    await bot.answer_callback_query(callback_query.id)
    await show_currency_options(callback_query.message, float(price), f"{duration}:{period}")

async def show_currency_options(message: types.Message, price: float, subscription_option: str):
    currencies = await get_available_currencies()
    keyboard = InlineKeyboardMarkup(row_width=2)
    for currency in currencies:
        converted_price = await convert_usd_to_crypto(price, currency)
        formatted_price = f"{converted_price:.8f}".rstrip('0').rstrip('.')
        keyboard.insert(InlineKeyboardButton(
            f"{currency} ({formatted_price})",
            callback_data=f"pay:{subscription_option}:{converted_price:.8f}:{currency}"
        ))
    await message.reply(f"Выберите валюту для оплаты (эквивалент {price}$):", reply_markup=keyboard)

async def convert_usd_to_crypto(usd_amount, crypto_currency):
    rates = {
        'BTC': 0.000020,
        'ETH': 0.00035,
        'USDT': 1,
        'TON': 0.14,
        'TRX': 8,
        'NOT': 85,
    }
    if crypto_currency in rates:
        return usd_amount * rates[crypto_currency]
    else:
        return usd_amount

async def show_main_menu(message: types.Message):
    user_id = message.from_user.id
    reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
        KeyboardButton('Начать рабочий день'),
        KeyboardButton('Скачать таблицу'),
        KeyboardButton('Аккаунт')
    )
    if user_id in ADMIN_IDS:
        reply_keyboard.add(KeyboardButton('Админ панель'))
    await message.reply("Выберите опцию из меню.", reply_markup=reply_keyboard)
    await UserState.WORKING_DAY_NOT_STARTED.set()
    
@dp.message_handler(Text(equals="Аккаунт"), state='*')
@check_active_subscription
async def account_info(message: types.Message, state: FSMContext, **kwargs):
    user_id = message.from_user.id
    if user_id in user_subscriptions:
        expiration_date = user_subscriptions[user_id]
        days_left = (expiration_date - datetime.datetime.now()).days
        
        # Здесь вы можете добавить код для получения статистики пользователя
        # Например, количество сделок, общая сумма и т.д.
        
        await message.reply(f"Ваша подписка действительна до: {expiration_date.strftime('%d.%m.%Y')}\n"
                            f"Осталось дней: {days_left}\n"
                            f"Статистика:\n"
                            f"- Количество сделок: X\n"
                            f"- Общая сумма сделок: Y RUB")
    else:
        await message.reply("У вас нет активной подписки.")
    
@dp.message_handler(commands=['admin'], state='*')
async def admin_panel(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.reply("У вас нет доступа к этой команде.")
        return
    
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("Просмотр подписок", callback_data="admin_view_subscriptions"))
    keyboard.add(InlineKeyboardButton("Выдать подписку", callback_data="admin_grant_subscription"))
    keyboard.add(InlineKeyboardButton("Выдать бесплатную подписку", callback_data="admin_grant_free_subscription"))
    
    await message.reply("Панель администратора:", reply_markup=keyboard)
    
@dp.callback_query_handler(lambda c: c.data == 'admin_grant_free_subscription')
async def grant_free_subscription_start(callback_query: types.CallbackQuery):
    if callback_query.from_user.id not in ADMIN_IDS:
        await bot.answer_callback_query(callback_query.id, "У вас нет доступа к этой функции.")
        return

    await bot.send_message(callback_query.from_user.id, "Введите ID пользователя и количество дней подписки в формате: ID дни")
    await bot.answer_callback_query(callback_query.id)
    await UserState.ADMIN_GRANTING_FREE_SUBSCRIPTION.set()
    
@dp.message_handler(state=UserState.ADMIN_GRANTING_FREE_SUBSCRIPTION)
async def grant_free_subscription(message: types.Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        await message.reply("У вас нет доступа к этой функции.")
        return

    try:
        user_id, days = map(int, message.text.split())
        expiration_date = datetime.datetime.now() + datetime.timedelta(days=days)
        user_subscriptions[user_id] = expiration_date
        save_subscriptions()
        
        # Сохраняем информацию о подписке в файл
        with open('subscriptions.txt', 'a') as f:
            f.write(f"{user_id},{expiration_date.isoformat()}\n")
        
        await message.reply(f"Бесплатная подписка выдана пользователю {user_id} до {expiration_date.strftime('%d.%m.%Y')}")
        
        # Отправляем уведомление пользователю
        await bot.send_message(user_id, f"Администратор системы выдал Вам бесплатную подписку на {days} дней, приятного пользования!")
    except ValueError:
        await message.reply("Неверный формат. Используйте: ID дни")
    
    await state.finish()
        
@dp.message_handler(Text(equals="Начать рабочий день"), state=UserState.WORKING_DAY_NOT_STARTED)
@check_active_subscription
async def start_day(message: types.Message, state: FSMContext, **kwargs):
    user_id = message.from_user.id
    file_path = os.path.join(USER_FILES_DIR, f'user_{user_id}.xlsx')
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ['Дата', 'Время', 'ID сделки', 'Тип сделки', 'Монета', 'Количество', 'Цена за единицу', 'Сумма в RUB', 'Способ оплаты', 'Реквизиты', 'Примечание']
        sheet.append(headers)
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

    now = datetime.datetime.now()
    date = now.strftime('%d.%m.%Y')
    time = now.strftime('%H:%M')
    row = [date, time, '', 'Начало рабочего дня', '', '', '', '', '', '', '']
    sheet.append(row)

    for col in range(1, 12):
        sheet.cell(row=sheet.max_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=11)
    sheet.cell(row=sheet.max_row, column=1).fill = PatternFill(start_color='FFFF00', fill_type='solid')

    workbook.save(file_path)

    reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
        KeyboardButton('Закончить рабочий день'),
        KeyboardButton('Скачать таблицу')
    )
    await message.reply('Рабочий день начат. Теперь вы можете добавлять сделки.', reply_markup=reply_keyboard)
    await state.set_state(UserState.WORKING_DAY_STARTED)

@dp.message_handler(Text(equals="Закончить рабочий день"), state=UserState.WORKING_DAY_STARTED)
@check_active_subscription
async def end_day(message: types.Message, state: FSMContext, **kwargs):
    user_id = message.from_user.id
    file_path = os.path.join(USER_FILES_DIR, f'user_{user_id}.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    now = datetime.datetime.now()
    date = now.strftime('%d.%m.%Y')
    time = now.strftime('%H:%M')
    row = [date, time, '', 'Конец рабочего дня', '', '', '', '', '', '', '']
    sheet.append(row)

    for col in range(1, 12):
        sheet.cell(row=sheet.max_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=11)
    sheet.cell(row=sheet.max_row, column=1).fill = PatternFill(start_color='FF0000', fill_type='solid')

    workbook.save(file_path)

    reply_keyboard = ReplyKeyboardMarkup(resize_keyboard=True).add(
        KeyboardButton('Начать рабочий день'),
        KeyboardButton('Скачать таблицу')
    )
    await message.reply('Рабочий день завершён. Добавление сделок временно приостановлено.', reply_markup=reply_keyboard)
    await state.set_state(UserState.WORKING_DAY_NOT_STARTED)

@dp.message_handler(Text(equals="Скачать таблицу"), state='*')
@check_active_subscription
async def download_table(message: types.Message, state: FSMContext, **kwargs):
    user_id = message.from_user.id
    file_path = os.path.join(USER_FILES_DIR, f'user_{user_id}.xlsx')
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            await bot.send_document(message.chat.id, types.InputFile(file, filename=f'user_{user_id}.xlsx'))
    else:
        await message.reply('У вас еще нет данных.')

def load_subscriptions():
    if os.path.exists('subscriptions.txt'):
        with open('subscriptions.txt', 'r') as f:
            for line in f:
                user_id, expiration_date_str = line.strip().split(',')
                user_subscriptions[int(user_id)] = datetime.datetime.fromisoformat(expiration_date_str)

# Вызовите эту функцию в on_startup
async def on_startup(_):
    await check_crypto_connection()
    asyncio.create_task(schedule_subscription_check(dp))
    load_subscriptions()
    async def on_shutdown(dp):
        save_subscriptions()
    
    dp.loop.add_signal_handler(signal.SIGINT, lambda: asyncio.create_task(on_shutdown(dp)))
    dp.loop.add_signal_handler(signal.SIGTERM, lambda: asyncio.create_task(on_shutdown(dp)))
    
def save_subscriptions():
    with open('subscriptions.txt', 'w') as f:
        for user_id, expiration_date in user_subscriptions.items():
            f.write(f"{user_id},{expiration_date.isoformat()}\n")
            
@dp.message_handler(state=UserState.WORKING_DAY_STARTED, content_types=types.ContentTypes.TEXT)
@check_active_subscription
async def add_data(message: types.Message, state: FSMContext, **kwargs):
    try:
        message_text = message.text
        user_id = message.from_user.id
        file_path = os.path.join(USER_FILES_DIR, f'user_{user_id}.xlsx')

        logger.info(f"Received message: {message_text}")

        if 'Сделка #' in message_text:
            try:
                # Extract information from the message text
                deal_id = re.search(r'(#\w+)', message_text).group(1)

                # Determine amount and coin
                amount_match = re.search(r'(Покупаете|Продаёте): ([\d,\.]+) (\w+)', message_text)
                amount = amount_match.group(2).replace(',', '')
                coin = amount_match.group(3)

                rub_amount_match = re.search(r'(Платите|Получаете): ([\d\s,\.]+) RUB', message_text)
                rub_amount = rub_amount_match.group(2).replace(' ', '').replace(',', '.') if rub_amount_match else ''

                price_per_coin_match = re.search(r'Цена за 1 \w+: ([\d,\.]+) RUB', message_text)
                price_per_coin = price_per_coin_match.group(1).replace(',', '.') if price_per_coin_match else ''

                payment_method = re.search(r'Способ оплаты: (.+)', message_text).group(1)

                requisites = re.search(r'Реквизиты: (.+)', message_text)
                requisites = requisites.group(1) if requisites else ''

                # Determine deal type
                deal_type = 'Покупка' if 'Покупаете' in message_text else 'Продажа'

                # Get deal time
                deal_time_match = re.search(r'✅ Сделка завершена (\d+) (\w+) (\d+), (\d+):(\d+):(\d+)', message_text)
                if deal_time_match:
                    day = int(deal_time_match.group(1))
                    month_str = deal_time_match.group(2)
                    year = int(deal_time_match.group(3))
                    hour = int(deal_time_match.group(4))
                    minute = int(deal_time_match.group(5))
                    second = int(deal_time_match.group(6))

                    # Convert month string to number
                    month_map = {
                        'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
                        'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
                    }
                    month = month_map.get(month_str.lower(), None)

                    if month is not None:
                        deal_time = datetime.datetime(year, month, day, hour, minute, second)

                        # Get date and time
                        date = deal_time.strftime('%d.%m.%Y')
                        time = deal_time.strftime('%H:%M')

                        # Open the Excel file and write data
                        workbook = openpyxl.load_workbook(file_path)
                        sheet = workbook.active

                        # Add new row
                        row = [date, time, deal_id, deal_type, coin, amount, price_per_coin, rub_amount,
                               payment_method, requisites, '']
                        sheet.append(row)

                        # Format cells
                        light_green_fill = PatternFill(start_color='90EE90', fill_type='solid')  # Light Green
                        light_red_fill = PatternFill(start_color='FFCCCC', fill_type='solid')  # Light Red
                        light_blue_fill = PatternFill(start_color='ADD8E6', fill_type='solid')  # Light Blue

                        if deal_type == 'Покупка':
                            sheet.cell(row=sheet.max_row, column=4).fill = light_green_fill
                        else:
                            sheet.cell(row=sheet.max_row, column=4).fill = light_red_fill

                        sheet.cell(row=sheet.max_row, column=5).fill = light_blue_fill

                        # Align data to center
                        for col in range(1, 12):
                            sheet.cell(row=sheet.max_row, column=col).alignment = Alignment(horizontal='center',
                                                                                            vertical='center')

                        workbook.save(file_path)

                        await message.reply('Данные о сделке успешно записаны в вашу таблицу Excel.')
                    else:
                        await message.reply('Не удалось распознать месяц в строке даты/времени. Пожалуйста, перешлите сообщение о сделке.')
                else:
                    await message.reply('Пожалуйста, перешлите сообщение о сделке, чтобы я мог записать информацию в таблицу Excel.')
            except Exception as e:
                logger.error(f"Error processing deal message: {str(e)}", exc_info=True)
                await message.reply(f'Ошибка при обработке сообщения: {e}')
        else:
            await message.reply('Пожалуйста, перешлите сообщение о сделке, чтобы записать информацию.')
    except Exception as e:
        logger.error(f"Error in add_data function: {str(e)}", exc_info=True)
        await message.reply(f'Ошибка при обработке сообщения: {e}')

@dp.message_handler(lambda message: message.from_user.id == CRYPTOBOT_USER_ID and "оплатил(а) ваш счёт" in message.text)
async def handle_cryptobot_payment(message: types.Message):
    logger.info(f"Received payment notification: {message.text}")
    invoice_id_match = re.search(r'#IV(\d+)', message.text)
    if invoice_id_match:
        invoice_id = invoice_id_match.group(1)
        for user_id, user_data in user_payments.items():
            if user_data['invoice_id'] == invoice_id:
                await process_successful_payment(user_id, user_data['amount'], user_data['subscription_option'])
                del user_payments[user_id]
                return
    logger.error(f"Could not process payment notification: {message.text}")

async def get_available_currencies():
    try:
        currencies = await crypto.get_currencies()
        supported_currencies = ['BTC', 'ETH', 'USDT', 'TON', 'TRX', 'NOT']
        return [currency.code for currency in currencies if currency.is_blockchain and currency.code in supported_currencies]
    except Exception as e:
        logger.error(f"Error getting available currencies: {str(e)}", exc_info=True)
        return ['USDT', 'TON', 'TRX', 'NOT']

async def on_startup(_):
    await check_crypto_connection()
    asyncio.create_task(schedule_subscription_check(dp))

if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)